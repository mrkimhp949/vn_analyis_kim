"""
Backtesting Engine - Kiểm tra hiệu suất chiến lược
"""

import concurrent.futures
import os
import sys
from datetime import datetime

# Add project root to Python path
project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if project_root not in sys.path:
    sys.path.insert(0, project_root)

import matplotlib
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd

from src.strategies.position_sizing import EnhancedPositionSizer
from src.data.loader import load_data
from src.ml.features.technical import add_ml_features
from src.ml.signals.generator import MLSignalGenerator

matplotlib.use("Agg")  # Sử dụng backend không cần GUI

# Fallback progress bar nếu không có tqdm
try:
    from tqdm import tqdm
except ImportError:
    print("⚠️ Tqdm not installed, using simple progress indicator")

    class tqdm:
        def __init__(self, iterable=None, total=None, desc=None):
            self.iterable = iterable
            self.total = total or len(iterable) if iterable else 0
            self.desc = desc or ""
            self.n = 0

        def __iter__(self):
            if self.iterable:
                for i, item in enumerate(self.iterable):
                    self.n = i + 1
                    if i % max(1, self.total // 10) == 0:  # Update every 10%
                        print(f"{self.desc}: {self.n}/{self.total} ({self.n/self.total*100:.1f}%)")
                    yield item
            else:
                for i in range(self.total):
                    self.n = i + 1
                    if i % max(1, self.total // 10) == 0:
                        print(f"{self.desc}: {self.n}/{self.total} ({self.n/self.total*100:.1f}%)")
                    yield i

        def update(self, n=1):
            self.n += n

        def close(self):
            pass

        def __enter__(self):
            return self

        def __exit__(self, *args):
            pass


class Backtester:
    def __init__(self, initial_capital=100_000_000, commission=0.0015, slippage=0.001):
        """
        initial_capital: Vốn ban đầu (VNĐ)
        commission: Phí giao dịch (0.15% mỗi chiều)
        slippage: Trượt giá 0.1%
        """
        self.initial_capital = initial_capital
        self.commission = commission
        self.slippage = slippage
        # lazy init ML generator may load models inside its constructor
        self.ml_generator = MLSignalGenerator()
        # Position sizer thận trọng
        self.position_sizer = EnhancedPositionSizer(
            total_capital=self.initial_capital,
            max_risk_per_trade=0.02,
            max_position_size=0.10,
            max_total_exposure=0.60,
        )

    def _apply_slippage(self, price, signal):
        """Áp dụng slippage cho giá"""
        slippage_factor = 1 + (self.slippage if signal == "BUY" else -self.slippage)
        return price * slippage_factor

    def run_backtest(
        self,
        symbol,
        start_date=None,
        end_date=None,
        lookback=500,
        confidence_threshold=50,
    ):
        """
        Chạy backtest trên 1 cổ phiếu với ML + Risk Management

        Args:
            confidence_threshold: Chỉ vào lệnh khi confidence >= threshold

        Returns:
            dict: Kết quả backtest với metrics
        """
        # Tạo thư mục backtest_results
        os.makedirs("backtest_results", exist_ok=True)

        print(f"\n{'='*60}")
        print(f"📊 BACKTESTING: {symbol}")
        print(f"💡 Confidence Threshold: {confidence_threshold}%")
        print(f"{'='*60}")

        # Load data - SỬA: thêm try-catch
        try:
            df = load_data(symbol, lookback=lookback)
            if df.empty:
                raise ValueError(f"Không có dữ liệu cho {symbol}")
        except (IOError, ValueError, KeyError, OSError) as e:
            print(f"❌ Lỗi load data {symbol}: {e}")
            return {
                "symbol": symbol,
                "initial_capital": self.initial_capital,
                "final_capital": self.initial_capital,
                "total_return": 0,
                "buy_hold_return": 0,
                "total_trades": 0,
                "winning_trades": 0,
                "losing_trades": 0,
                "win_rate": 0,
                "max_drawdown": 0,
                "sharpe_ratio": 0,
                "avg_confidence": 0,
                "confidence_threshold": confidence_threshold,
                "trades": pd.DataFrame(),
                "portfolio_values": pd.DataFrame(),
            }

        if start_date:
            df = df[df["time"] >= start_date]
        if end_date:
            df = df[df["time"] <= end_date]

        if df.empty:
            print(f"❌ Không có dữ liệu sau filter cho {symbol}")
            return {
                "symbol": symbol,
                "initial_capital": self.initial_capital,
                "final_capital": self.initial_capital,
                "total_return": 0,
                "buy_hold_return": 0,
                "total_trades": 0,
                "winning_trades": 0,
                "losing_trades": 0,
                "win_rate": 0,
                "max_drawdown": 0,
                "sharpe_ratio": 0,
                "avg_confidence": 0,
                "confidence_threshold": confidence_threshold,
                "trades": pd.DataFrame(),
                "portfolio_values": pd.DataFrame(),
            }

        start_range = df["time"].min().date()
        end_range = df["time"].max().date()
        print(f"📅 Từ {start_range} đến {end_range}")
        print(f"📈 Tổng số ngày: {len(df)}")

        # Initialize ML (use lazy-initialized self.ml_generator)
        ml_generator = self.ml_generator

        # Portfolio
        capital = self.initial_capital
        position = 0  # Số cổ phiếu đang nắm giữ
        trades = []
        portfolio_values = []
        position_data = None

        # Simulate trading với progress indicator
        total_days = len(df) - 50
        if total_days <= 0:
            print("❌ Không đủ dữ liệu để backtest (cần >50 ngày)")
            return {
                "symbol": symbol,
                "initial_capital": self.initial_capital,
                "final_capital": self.initial_capital,
                "total_return": 0,
                "buy_hold_return": 0,
                "total_trades": 0,
                "winning_trades": 0,
                "losing_trades": 0,
                "win_rate": 0,
                "max_drawdown": 0,
                "sharpe_ratio": 0,
                "avg_confidence": 0,
                "confidence_threshold": confidence_threshold,
                "trades": pd.DataFrame(),
                "portfolio_values": pd.DataFrame(),
            }

        print(f"⏳ Đang mô phỏng {total_days} ngày giao dịch...")

        for i in range(50, len(df)):  # Skip first 50 days for indicators
            current_data = df.iloc[: i + 1].copy()
            current_row = df.iloc[i]

            # Hiển thị progress mỗi 10%
            if (i - 50) % max(1, total_days // 10) == 0:
                progress = ((i - 50) / total_days) * 100
                print(f"  📊 Đang xử lý: {progress:.1f}%")

            # ML Analysis
            try:
                result = ml_generator.analyze(current_data)
                signal = result.get("signal", "HOLD")
                confidence = result.get("confidence", 0)
                price = current_row["close"]
                day_high = current_row.get("high", price)
                day_low = current_row.get("low", price)
                # Calculate ATR with fallback for insufficient data
                if "atr" in current_row and pd.notna(current_row["atr"]):
                    atr_value = current_row["atr"]
                elif len(current_data) >= 14:
                    atr_value = current_data["close"].rolling(14).std().iloc[-1]
                    atr_value = atr_value if pd.notna(atr_value) else price * 0.02
                else:
                    atr_value = price * 0.02  # Default to 2% of price if insufficient data

                # Kiểm tra stop-loss / take-profit khi đang giữ vị thế
                if position > 0 and position_data:
                    exit_reason = None
                    exit_price = None
                    if day_low <= position_data["stop_loss"]:
                        exit_price = position_data["stop_loss"]
                        exit_reason = "STOP_LOSS"
                    elif day_high >= position_data["take_profit"]:
                        exit_price = position_data["take_profit"]
                        exit_reason = "TAKE_PROFIT"

                    if exit_reason and exit_price:
                        revenue = position * exit_price * (1 - self.commission)
                        capital += revenue
                        trades.append(
                            {
                                "date": current_row["time"],
                                "type": "SELL",
                                "price": exit_price,
                                "shares": position,
                                "value": revenue,
                                "confidence": confidence,
                                "ml_score": result.get("ml_score", 0),
                                "exit_reason": exit_reason,
                            }
                        )
                        position = 0
                        position_data = None
                        # Sau khi chốt, không tiếp tục xử lý SIGNAL SELL cùng ngày
                        if signal != "BUY":
                            continue

                # Áp dụng slippage
                execution_price = self._apply_slippage(price, signal)

                # Validate confidence is a valid number
                if not isinstance(confidence, (int, float)) or pd.isna(confidence):
                    confidence = 0  # Default to 0 if invalid

                # CHỈ VÀO LỆNH KHI CONFIDENCE >= THRESHOLD
                if signal == "BUY" and confidence >= confidence_threshold and position == 0:
                    # Tính ATR và stop loss dựa trên features
                    enriched = add_ml_features(current_data)
                    latest_atr = float(enriched.iloc[-1].get("atr", 0)) if len(enriched) > 0 else 0
                    atr_for_stop = (
                        latest_atr
                        if latest_atr > 0
                        else (
                            atr_value
                            if pd.notna(atr_value) and atr_value > 0
                            else execution_price * 0.03
                        )
                    )
                    stop_loss_price = max(0, execution_price - 2.0 * atr_for_stop)
                    take_profit_price = execution_price + 3.0 * atr_for_stop

                    # Sizing an toàn bằng ConservativePositionSizer
                    sized = self.position_sizer.calculate_position_size(
                        symbol=symbol,
                        entry_price=execution_price,
                        stop_loss=stop_loss_price,
                        confidence=int(confidence),
                        signal_strength=(
                            "VERY_STRONG"
                            if confidence >= 80
                            else ("STRONG" if confidence >= 65 else "MODERATE")
                        ),
                        market_regime={"regime": "SIDEWAYS", "tradeable": True},
                    )
                    shares_to_buy = int(sized.shares)
                    cost = shares_to_buy * execution_price * (1 + self.commission)

                    if cost <= capital and shares_to_buy > 0:
                        position = shares_to_buy
                        capital -= cost

                        trades.append(
                            {
                                "date": current_row["time"],
                                "type": "BUY",
                                "price": execution_price,
                                "shares": shares_to_buy,
                                "value": cost,
                                "confidence": confidence,
                                "ml_score": result.get("ml_score", 0),
                                "atr": latest_atr,
                                "stop_loss": stop_loss_price,
                                "exit_reason": "OPEN",
                            }
                        )
                        position_data = {
                            "entry_price": execution_price,
                            "stop_loss": stop_loss_price,
                            "take_profit": take_profit_price,
                        }

                elif signal == "SELL" and confidence >= confidence_threshold and position > 0:
                    # Bán
                    revenue = position * execution_price * (1 - self.commission)
                    capital += revenue

                    trades.append(
                        {
                            "date": current_row["time"],
                            "type": "SELL",
                            "price": execution_price,
                            "shares": position,
                            "value": revenue,
                            "confidence": confidence,
                            "ml_score": result.get("ml_score", 0),
                            "exit_reason": "SIGNAL_SELL",
                        }
                    )

                    position = 0
                    position_data = None

                # Portfolio value
                portfolio_value = capital + (position * price if position > 0 else 0)
                portfolio_values.append(
                    {
                        "date": current_row["time"],
                        "value": portfolio_value,
                        "price": price,
                    }
                )

            except (KeyError, ValueError, AttributeError, IndexError, TypeError) as e:
                print(f"⚠️ Lỗi ngày {current_row['time'].date()}: {e}")

        # Close any open position
        if position > 0:
            final_price = df.iloc[-1]["close"]
            final_execution_price = self._apply_slippage(final_price, "SELL")
            capital += position * final_execution_price * (1 - self.commission)
            trades.append(
                {
                    "date": df.iloc[-1]["time"],
                    "type": "SELL",
                    "price": final_execution_price,
                    "shares": position,
                    "value": position * final_execution_price,
                    "confidence": 0,
                    "ml_score": 0,
                    "exit_reason": "EOD_EXIT",
                }
            )
            position = 0
            position_data = None

        # Calculate metrics
        final_capital = capital
        total_return = (final_capital - self.initial_capital) / self.initial_capital * 100

        # Buy & Hold comparison (with division by zero protection)
        start_price = df.iloc[50]["close"] if len(df) > 50 else df.iloc[0]["close"]
        buy_hold_return = (
            ((df.iloc[-1]["close"] - start_price) / start_price) * 100 if start_price > 0 else 0
        )

        # Analyze trades
        trades_df = pd.DataFrame(trades)
        winning_trades = 0
        losing_trades = 0
        trade_profits = []
        gross_profit = 0.0
        gross_loss = 0.0
        consecutive_losses = 0
        max_consecutive_losses = 0

        if len(trades_df) > 0:
            # Properly pair BUY-SELL trades by time order (FIFO)
            trade_pairs = []
            open_buys = []  # Stack of open BUY positions

            for idx, trade in trades_df.iterrows():
                if trade["type"] == "BUY":
                    open_buys.append(trade)
                elif trade["type"] == "SELL" and open_buys:
                    buy_trade = open_buys.pop(0)  # FIFO: first in, first out
                    trade_pairs.append((buy_trade, trade))

            # Calculate P&L for each properly paired trade
            for buy_trade, sell_trade in trade_pairs:
                buy_price = buy_trade["price"]
                sell_price = sell_trade["price"]
                shares_traded = sell_trade["shares"]
                pnl = (
                    (sell_price - buy_price) * shares_traded
                    - (buy_price * shares_traded * self.commission)
                    - (sell_price * shares_traded * self.commission)
                )
                trade_profits.append(pnl)

                if pnl > 0:
                    winning_trades += 1
                    gross_profit += pnl
                    consecutive_losses = 0
                else:
                    losing_trades += 1
                    gross_loss += abs(pnl)
                    consecutive_losses += 1
                    max_consecutive_losses = max(max_consecutive_losses, consecutive_losses)

        win_rate = (
            (winning_trades / (winning_trades + losing_trades) * 100)
            if (winning_trades + losing_trades) > 0
            else 0
        )

        # Portfolio values
        portfolio_df = pd.DataFrame(portfolio_values)
        max_drawdown = (
            self._calculate_max_drawdown(portfolio_df["value"].values)
            if not portfolio_df.empty
            else 0
        )

        # Sharpe Ratio
        returns = portfolio_df["value"].pct_change().dropna()
        sharpe_ratio = (
            (returns.mean() / returns.std() * np.sqrt(252))
            if len(returns) > 0 and returns.std() > 0
            else 0
        )

        downside_returns = returns[returns < 0]
        sortino_ratio = (
            (returns.mean() / downside_returns.std() * np.sqrt(252))
            if len(downside_returns) > 0 and downside_returns.std() > 0
            else 0
        )
        annual_return = (
            ((final_capital / self.initial_capital) ** (252 / max(len(portfolio_df), 1)) - 1)
            if len(portfolio_df) > 0
            else 0
        )
        calmar_ratio = (annual_return / (max_drawdown / 100)) if max_drawdown > 0 else 0
        # Profit factor: if no losses, return gross_profit (or 0 if no profit either)
        profit_factor = (
            (gross_profit / gross_loss)
            if gross_loss > 0
            else (gross_profit if gross_profit > 0 else 0)
        )

        # Average confidence
        avg_confidence = trades_df["confidence"].mean() if len(trades_df) > 0 else 0

        results = {
            "symbol": symbol,
            "initial_capital": self.initial_capital,
            "final_capital": final_capital,
            "total_return": total_return,
            "buy_hold_return": buy_hold_return,
            "total_trades": len(trades),
            "winning_trades": winning_trades,
            "losing_trades": losing_trades,
            "win_rate": win_rate,
            "max_drawdown": max_drawdown,
            "sharpe_ratio": sharpe_ratio,
            "sortino_ratio": sortino_ratio,
            "calmar_ratio": calmar_ratio,
            "profit_factor": profit_factor,
            "max_consecutive_losses": max_consecutive_losses,
            "avg_confidence": avg_confidence,
            "confidence_threshold": confidence_threshold,
            "trades": trades_df,
            "portfolio_values": portfolio_df,
        }

        self._print_results(results)

        return results

    def run_multiple_backtest_parallel(
        self, symbols, lookback=500, confidence_threshold=50, max_workers=3
    ):
        """Chạy backtest song song cho nhiều cổ phiếu"""
        print(f"🚀 Chạy backtest song song {len(symbols)} mã (workers: {max_workers})")

        def run_single(symbol):
            try:
                return self.run_backtest(
                    symbol, lookback=lookback, confidence_threshold=confidence_threshold
                )
            except (KeyError, ValueError, IOError, OSError, RuntimeError) as e:
                print(f"❌ Lỗi {symbol}: {e}")
                return None

        with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
            results = list(
                tqdm(
                    executor.map(run_single, symbols),
                    total=len(symbols),
                    desc="Backtesting",
                )
            )

        # Lọc kết quả None
        valid_results = [r for r in results if r is not None]
        self._print_summary(valid_results)

        return valid_results

    def _calculate_max_drawdown(self, values):
        """Tính Max Drawdown"""
        if len(values) == 0:
            return 0

        peak = values[0]
        max_dd = 0

        for value in values:
            if value > peak:
                peak = value
            dd = (peak - value) / peak * 100
            if dd > max_dd:
                max_dd = dd

        return max_dd

    def _print_results(self, results):
        """In kết quả backtest"""
        print("\n📈 KẾT QUẢ BACKTEST")
        print(f"{'='*60}")
        print(f"💰 Vốn ban đầu:        {results['initial_capital']:>15,} VNĐ")
        print(f"💵 Vốn cuối:           {results['final_capital']:>15,.0f} VNĐ")
        print(f"📊 Lợi nhuận:          {results['total_return']:>14.2f} %")
        print(f"🎯 Buy & Hold:         {results['buy_hold_return']:>14.2f} %")
        print(f"🔄 Tổng giao dịch:     {results['total_trades']:>15}")
        print(f"✅ Thắng:              {results['winning_trades']:>15}")
        print(f"❌ Thua:               {results['losing_trades']:>15}")
        print(f"🎲 Tỷ lệ thắng:        {results['win_rate']:>14.2f} %")
        print(f"📉 Max Drawdown:       {results['max_drawdown']:>14.2f} %")
        print(f"📈 Sharpe Ratio:       {results['sharpe_ratio']:>14.2f}")
        print(f"📈 Sortino Ratio:      {results['sortino_ratio']:>14.2f}")
        print(f"📈 Calmar Ratio:       {results['calmar_ratio']:>14.2f}")
        print(f"💹 Profit Factor:      {results['profit_factor']:>14.2f}")
        print(f"⚠️ Chuỗi thua tối đa:  {results['max_consecutive_losses']:>15}")
        print(f"{'='*60}\n")

        # Performance vs Buy&Hold
        outperformance = results["total_return"] - results["buy_hold_return"]
        if outperformance > 0:
            print(f"🚀 Chiến lược VƯỢT QUÁ Buy&Hold: +{outperformance:.2f}%")
        else:
            print(f"📉 Chiến lược KÉMHƠN Buy&Hold: {outperformance:.2f}%")

    def plot_results(self, results):
        """Vẽ biểu đồ kết quả"""
        # Tạo thư mục backtest_results nếu chưa có
        try:
            os.makedirs("backtest_results", exist_ok=True)
            print(f"📁 Thư mục backtest_results: {os.path.abspath('backtest_results')}")
        except OSError as e:
            print(f"❌ Lỗi tạo thư mục: {e}")
            return

        portfolio_df = results["portfolio_values"]
        trades_df = results["trades"]

        if portfolio_df.empty:
            print("❌ Không có dữ liệu portfolio để vẽ biểu đồ")
            return

        fig, axes = plt.subplots(2, 1, figsize=(14, 10))

        # Portfolio value
        axes[0].plot(
            portfolio_df["date"],
            portfolio_df["value"],
            label="Portfolio Value",
            linewidth=2,
            color="#2E86AB",
        )
        axes[0].axhline(
            y=self.initial_capital,
            color="gray",
            linestyle="--",
            label="Initial Capital",
            alpha=0.7,
        )
        axes[0].fill_between(
            portfolio_df["date"],
            self.initial_capital,
            portfolio_df["value"],
            where=(portfolio_df["value"] >= self.initial_capital),
            color="green",
            alpha=0.1,
            label="Profit",
        )
        axes[0].fill_between(
            portfolio_df["date"],
            self.initial_capital,
            portfolio_df["value"],
            where=(portfolio_df["value"] < self.initial_capital),
            color="red",
            alpha=0.1,
            label="Loss",
        )
        axes[0].set_title(
            f"{results['symbol']} - Portfolio Value Over Time",
            fontsize=14,
            fontweight="bold",
        )
        axes[0].set_ylabel("Value (VNĐ)", fontsize=12)
        axes[0].legend(loc="best")
        axes[0].grid(True, alpha=0.3)
        axes[0].yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f"{x:,.0f}"))

        # Price + Buy/Sell signals
        axes[1].plot(
            portfolio_df["date"],
            portfolio_df["price"],
            label="Price",
            linewidth=1.5,
            alpha=0.7,
            color="#A23B72",
        )

        if len(trades_df) > 0:
            buy_trades = trades_df[trades_df["type"] == "BUY"]
            sell_trades = trades_df[trades_df["type"] == "SELL"]

            if len(buy_trades) > 0:
                axes[1].scatter(
                    buy_trades["date"],
                    buy_trades["price"],
                    color="green",
                    marker="^",
                    s=150,
                    label="BUY",
                    zorder=5,
                    edgecolors="darkgreen",
                    linewidth=1.5,
                )
            if len(sell_trades) > 0:
                axes[1].scatter(
                    sell_trades["date"],
                    sell_trades["price"],
                    color="red",
                    marker="v",
                    s=150,
                    label="SELL",
                    zorder=5,
                    edgecolors="darkred",
                    linewidth=1.5,
                )

        axes[1].set_title(
            f"{results['symbol']} - Price & Trading Signals",
            fontsize=14,
            fontweight="bold",
        )
        axes[1].set_xlabel("Date", fontsize=12)
        axes[1].set_ylabel("Price (VNĐ)", fontsize=12)
        axes[1].legend(loc="best")
        axes[1].grid(True, alpha=0.3)
        axes[1].yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f"{x:,.0f}"))

        # Add text with key metrics
        fig.text(
            0.99,
            0.01,
            f"Return: {results['total_return']:.2f}% | "
            f"Win Rate: {results['win_rate']:.1f}% | Sharpe: {results['sharpe_ratio']:.2f}",
            ha="right",
            va="bottom",
            fontsize=10,
            style="italic",
            alpha=0.7,
        )

        plt.tight_layout()

        # Lưu với timestamp
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f'backtest_results/{results["symbol"]}_{timestamp}.png'

            print(f"💾 Đang lưu biểu đồ vào: {filename}")
            plt.savefig(filename, dpi=300, bbox_inches="tight")

            # Kiểm tra file
            if os.path.exists(filename):
                file_size = os.path.getsize(filename)
                print(f"✅ Biểu đồ đã lưu: {filename} ({file_size} bytes)")
            else:
                print("❌ Không tìm thấy file sau khi lưu!")

            plt.show()

        except (IOError, OSError, KeyError, ValueError) as e:
            print(f"❌ Lỗi khi lưu biểu đồ: {e}")
            import traceback

            traceback.print_exc()

    def run_multiple_backtest(self, symbols, lookback=500, confidence_threshold=50):
        """Chạy backtest cho nhiều cổ phiếu
        Thêm tham số confidence_threshold để đồng bộ với giao diện gọi."""
        all_results = []

        for symbol in symbols:
            try:
                result = self.run_backtest(
                    symbol, lookback=lookback, confidence_threshold=confidence_threshold
                )
                all_results.append(result)
            except (KeyError, ValueError, IOError, OSError, RuntimeError) as e:
                print(f"❌ Lỗi backtest {symbol}: {e}")

        # Summary
        self._print_summary(all_results)

        return all_results

    # Trong method _print_summary, sửa phần export CSV:
    def _print_summary(self, all_results):
        """Tổng kết backtest nhiều cổ phiếu"""
        print(f"\n{'='*60}")
        print("📊 TỔNG KẾT BACKTEST")
        print(f"{'='*60}")

        if not all_results:
            print("❌ Không có kết quả nào để tổng kết")
            return

        summary_df = pd.DataFrame(
            [
                {
                    "Symbol": r["symbol"],
                    "Return (%)": f"{r['total_return']:.2f}",
                    "Buy&Hold (%)": f"{r['buy_hold_return']:.2f}",
                    "Trades": r["total_trades"],
                    "Win Rate (%)": f"{r['win_rate']:.2f}",
                    "Sharpe": f"{r['sharpe_ratio']:.2f}",
                    "Sortino": f"{r['sortino_ratio']:.2f}",
                    "Calmar": f"{r['calmar_ratio']:.2f}",
                    "Profit Factor": f"{r['profit_factor']:.2f}",
                    "Max Loss Streak": r["max_consecutive_losses"],
                }
                for r in all_results
            ]
        )

        print(summary_df.to_string(index=False))
        print(f"{'='*60}\n")

        # Tự động lưu summary với Excel
        try:
            os.makedirs("backtest_results", exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

            # DataFrame chi tiết cho export
            detailed_summary = pd.DataFrame(
                [
                    {
                        "Symbol": r["symbol"],
                        "Initial Capital": r["initial_capital"],
                        "Final Capital": r["final_capital"],
                        "Return (%)": r["total_return"],
                        "Buy&Hold (%)": r["buy_hold_return"],
                        "Outperformance (%)": r["total_return"] - r["buy_hold_return"],
                        "Total Trades": r["total_trades"],
                        "Winning Trades": r["winning_trades"],
                        "Losing Trades": r["losing_trades"],
                        "Win Rate (%)": r["win_rate"],
                        "Max Drawdown (%)": r["max_drawdown"],
                        "Sharpe Ratio": r["sharpe_ratio"],
                        "Sortino Ratio": r["sortino_ratio"],
                        "Calmar Ratio": r["calmar_ratio"],
                        "Profit Factor": r["profit_factor"],
                        "Max Consecutive Losses": r["max_consecutive_losses"],
                    }
                    for r in all_results
                ]
            )

            # Export Excel với format đẹp
            excel_filename = f"backtest_results/summary_{timestamp}.xlsx"

            with pd.ExcelWriter(excel_filename, engine="openpyxl") as writer:
                detailed_summary.to_excel(writer, sheet_name="Summary", index=False)

                # Format worksheet
                worksheet = writer.sheets["Summary"]

                # Auto-adjust column width
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except (AttributeError, TypeError):
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

                # Header style
                from openpyxl.styles import Alignment, Font, PatternFill

                header_fill = PatternFill(
                    start_color="366092", end_color="366092", fill_type="solid"
                )
                header_font = Font(color="FFFFFF", bold=True)

                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                # Number format
                for row in worksheet.iter_rows(min_row=2):
                    # Return, Buy&Hold, Outperformance columns
                    for idx in [3, 4, 5, 9, 10]:
                        row[idx].number_format = "0.00"
                    # Capital columns
                    for idx in [1, 2]:
                        row[idx].number_format = "#,##0"

                # Conditional formatting cho Return
                from openpyxl.formatting.rule import CellIsRule

                green_fill = PatternFill(
                    start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
                )
                red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

                worksheet.conditional_formatting.add(
                    f"D2:D{len(detailed_summary)+1}",
                    CellIsRule(operator="greaterThan", formula=["0"], fill=green_fill),
                )
                worksheet.conditional_formatting.add(
                    f"D2:D{len(detailed_summary)+1}",
                    CellIsRule(operator="lessThan", formula=["0"], fill=red_fill),
                )

            if os.path.exists(excel_filename):
                file_size = os.path.getsize(excel_filename)
                print(f"✅ Đã xuất Excel: {excel_filename} ({file_size:,} bytes)")

            # Cũng export CSV đơn giản - SỬA LỖI ENCODING
            csv_filename = f"backtest_results/summary_{timestamp}.csv"
            try:
                detailed_summary.to_csv(csv_filename, index=False, encoding="utf-8-sig")
            except UnicodeEncodeError:
                # Fallback cho Windows
                detailed_summary.to_csv(csv_filename, index=False, encoding="cp1252")
            print(f"✅ Đã xuất CSV: {csv_filename}\n")

        except ImportError:
            # Fallback nếu không có openpyxl
            print("⚠️ Cài đặt openpyxl để xuất Excel đẹp: pip install openpyxl")
            csv_filename = f"backtest_results/summary_{timestamp}.csv"
            try:
                detailed_summary.to_csv(csv_filename, index=False, encoding="utf-8-sig")
            except UnicodeEncodeError:
                detailed_summary.to_csv(csv_filename, index=False, encoding="cp1252")
            print(f"✅ Đã xuất CSV: {csv_filename}\n")

        except (IOError, OSError, KeyError, ValueError) as e:
            print(f"❌ Lỗi khi xuất file: {e}\n")
            import traceback

            traceback.print_exc()
