"""
Backtesting Engine - Kiểm tra hiệu suất chiến lược
"""

import pandas as pd
import numpy as np
from datetime import datetime
import matplotlib.pyplot as plt
from data_loader import load_data
from ml_signals import MLSignalGenerator

class Backtester:
    def __init__(self, initial_capital=100_000_000, commission=0.0015):
        """
        initial_capital: Vốn ban đầu (VNĐ)
        commission: Phí giao dịch (0.15% mỗi chiều)
        """
        self.initial_capital = initial_capital
        self.commission = commission
        # lazy init ML generator may load models inside its constructor
        self.ml_generator = MLSignalGenerator()
        
    def run_backtest(self, symbol, start_date=None, end_date=None, lookback=500, confidence_threshold=50):
        """
        Chạy backtest trên 1 cổ phiếu với ML + Risk Management
        
        Args:
            confidence_threshold: Chỉ vào lệnh khi confidence >= threshold
        
        Returns:
            dict: Kết quả backtest với metrics
        """
        # Tạo thư mục backtest_results
        import os
        os.makedirs('backtest_results', exist_ok=True)
        
        print(f"\n{'='*60}")
        print(f"📊 BACKTESTING: {symbol}")
        print(f"💡 Confidence Threshold: {confidence_threshold}%")
        print(f"{'='*60}")
        
        # Load data
        df = load_data(symbol, lookback=lookback)
        
        if start_date:
            df = df[df['time'] >= start_date]
        if end_date:
            df = df[df['time'] <= end_date]
        
        print(f"📅 Từ {df['time'].min().date()} đến {df['time'].max().date()}")
        print(f"📈 Tổng số ngày: {len(df)}")
        
        # Initialize ML (use lazy-initialized self.ml_generator)
        ml_generator = self.ml_generator
        
        # Portfolio
        capital = self.initial_capital
        position = 0  # Số cổ phiếu đang nắm giữ
        trades = []
        portfolio_values = []
        
        # Simulate trading
        for i in range(50, len(df)):  # Skip first 50 days for indicators
            current_data = df.iloc[:i+1].copy()
            current_row = df.iloc[i]
            
            # ML Analysis
            try:
                result = ml_generator.analyze(current_data)
                signal = result.get('signal', 'HOLD')
                confidence = result.get('confidence', 0)
                price = current_row['close']
                
                # CHỈ VÀO LỆNH KHI CONFIDENCE >= THRESHOLD
                if signal == 'BUY' and confidence >= confidence_threshold and position == 0:
                    # Mua
                    shares_to_buy = int(capital * 0.95 / price)  # Dùng 95% vốn
                    cost = shares_to_buy * price * (1 + self.commission)
                    
                    if cost <= capital and shares_to_buy > 0:
                        position = shares_to_buy
                        capital -= cost
                        
                        trades.append({
                            'date': current_row['time'],
                            'type': 'BUY',
                            'price': price,
                            'shares': shares_to_buy,
                            'value': cost,
                            'confidence': confidence,
                            'ml_score': result.get('ml_score', 0)
                        })
                
                elif signal == 'SELL' and confidence >= confidence_threshold and position > 0:
                    # Bán
                    revenue = position * price * (1 - self.commission)
                    capital += revenue
                    
                    trades.append({
                        'date': current_row['time'],
                        'type': 'SELL',
                        'price': price,
                        'shares': position,
                        'value': revenue,
                        'confidence': confidence,
                        'ml_score': result.get('ml_score', 0)
                    })
                    
                    position = 0
                
                # Portfolio value
                portfolio_value = capital + (position * price if position > 0 else 0)
                portfolio_values.append({
                    'date': current_row['time'],
                    'value': portfolio_value,
                    'price': price
                })
                
            except Exception as e:
                print(f"⚠️ Lỗi ngày {current_row['time'].date()}: {e}")
        
        # Close any open position
        if position > 0:
            final_price = df.iloc[-1]['close']
            capital += position * final_price * (1 - self.commission)
            trades.append({
                'date': df.iloc[-1]['time'],
                'type': 'SELL',
                'price': final_price,
                'shares': position,
                'value': position * final_price,
                'confidence': 0,
                'ml_score': 0
            })
            position = 0
        
        # Calculate metrics
        final_capital = capital
        total_return = (final_capital - self.initial_capital) / self.initial_capital * 100
        
        # Buy & Hold comparison
        buy_hold_return = ((df.iloc[-1]['close'] - df.iloc[50]['close']) / df.iloc[50]['close']) * 100
        
        # Analyze trades
        trades_df = pd.DataFrame(trades)
        winning_trades = 0
        losing_trades = 0
        
        if len(trades_df) > 0:
            buy_trades = trades_df[trades_df['type'] == 'BUY']
            sell_trades = trades_df[trades_df['type'] == 'SELL']
            
            for i in range(min(len(buy_trades), len(sell_trades))):
                buy_price = buy_trades.iloc[i]['price']
                sell_price = sell_trades.iloc[i]['price']
                
                if sell_price > buy_price:
                    winning_trades += 1
                else:
                    losing_trades += 1
        
        win_rate = (winning_trades / (winning_trades + losing_trades) * 100) if (winning_trades + losing_trades) > 0 else 0
        
        # Portfolio values
        portfolio_df = pd.DataFrame(portfolio_values)
        max_drawdown = self._calculate_max_drawdown(portfolio_df['value'].values) if not portfolio_df.empty else 0
        
        # Sharpe Ratio
        returns = portfolio_df['value'].pct_change().dropna()
        sharpe_ratio = (returns.mean() / returns.std() * np.sqrt(252)) if len(returns) > 0 and returns.std() > 0 else 0
        
        # Average confidence
        avg_confidence = trades_df['confidence'].mean() if len(trades_df) > 0 else 0
        
        results = {
            'symbol': symbol,
            'initial_capital': self.initial_capital,
            'final_capital': final_capital,
            'total_return': total_return,
            'buy_hold_return': buy_hold_return,
            'total_trades': len(trades),
            'winning_trades': winning_trades,
            'losing_trades': losing_trades,
            'win_rate': win_rate,
            'max_drawdown': max_drawdown,
            'sharpe_ratio': sharpe_ratio,
            'avg_confidence': avg_confidence,
            'confidence_threshold': confidence_threshold,
            'trades': trades_df,
            'portfolio_values': portfolio_df
        }
        
        self._print_results(results)
        
        return results
    
    def _calculate_max_drawdown(self, values):
        """Tính Max Drawdown"""
        peak = values[0]
        max_dd = 0
        
        for value in values:
            if value > peak:
                peak = value
            dd = (peak - value) / peak * 100
            if dd > max_dd:
                max_dd = dd
        
        return max_dd
    
    def _print_results(self, results):
        """In kết quả backtest"""
        print(f"\n📈 KẾT QUẢ BACKTEST")
        print(f"{'='*60}")
        print(f"💰 Vốn ban đầu:        {results['initial_capital']:>15,} VNĐ")
        print(f"💵 Vốn cuối:           {results['final_capital']:>15,.0f} VNĐ")
        print(f"📊 Lợi nhuận:          {results['total_return']:>14.2f} %")
        print(f"🎯 Buy & Hold:         {results['buy_hold_return']:>14.2f} %")
        print(f"🔄 Tổng giao dịch:     {results['total_trades']:>15}")
        print(f"✅ Thắng:              {results['winning_trades']:>15}")
        print(f"❌ Thua:               {results['losing_trades']:>15}")
        print(f"🎲 Tỷ lệ thắng:        {results['win_rate']:>14.2f} %")
        print(f"📉 Max Drawdown:       {results['max_drawdown']:>14.2f} %")
        print(f"📈 Sharpe Ratio:       {results['sharpe_ratio']:>14.2f}")
        print(f"{'='*60}\n")
        
        # Performance vs Buy&Hold
        outperformance = results['total_return'] - results['buy_hold_return']
        if outperformance > 0:
            print(f"🚀 Chiến lược VƯỢT QUÁ Buy&Hold: +{outperformance:.2f}%")
        else:
            print(f"📉 Chiến lược KÉMHƠN Buy&Hold: {outperformance:.2f}%")
    
    def plot_results(self, results):
        """Vẽ biểu đồ kết quả"""
        import os
        from datetime import datetime
        
        # Tạo thư mục backtest_results nếu chưa có
        try:
            os.makedirs('backtest_results', exist_ok=True)
            print(f"📁 Thư mục backtest_results: {os.path.abspath('backtest_results')}")
        except Exception as e:
            print(f"❌ Lỗi tạo thư mục: {e}")
            return
        
        portfolio_df = results['portfolio_values']
        trades_df = results['trades']
        
        fig, axes = plt.subplots(2, 1, figsize=(14, 10))
        
        # Portfolio value
        axes[0].plot(portfolio_df['date'], portfolio_df['value'], label='Portfolio Value', linewidth=2, color='#2E86AB')
        axes[0].axhline(y=self.initial_capital, color='gray', linestyle='--', label='Initial Capital', alpha=0.7)
        axes[0].fill_between(portfolio_df['date'], self.initial_capital, portfolio_df['value'], 
                            where=(portfolio_df['value'] >= self.initial_capital), 
                            color='green', alpha=0.1, label='Profit')
        axes[0].fill_between(portfolio_df['date'], self.initial_capital, portfolio_df['value'], 
                            where=(portfolio_df['value'] < self.initial_capital), 
                            color='red', alpha=0.1, label='Loss')
        axes[0].set_title(f"{results['symbol']} - Portfolio Value Over Time", fontsize=14, fontweight='bold')
        axes[0].set_ylabel('Value (VNĐ)', fontsize=12)
        axes[0].legend(loc='best')
        axes[0].grid(True, alpha=0.3)
        axes[0].yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'{x:,.0f}'))
        
        # Price + Buy/Sell signals
        axes[1].plot(portfolio_df['date'], portfolio_df['price'], label='Price', linewidth=1.5, alpha=0.7, color='#A23B72')
        
        if len(trades_df) > 0:
            buy_trades = trades_df[trades_df['type'] == 'BUY']
            sell_trades = trades_df[trades_df['type'] == 'SELL']
            
            if len(buy_trades) > 0:
                axes[1].scatter(buy_trades['date'], buy_trades['price'], 
                              color='green', marker='^', s=150, label='BUY', zorder=5, edgecolors='darkgreen', linewidth=1.5)
            if len(sell_trades) > 0:
                axes[1].scatter(sell_trades['date'], sell_trades['price'], 
                              color='red', marker='v', s=150, label='SELL', zorder=5, edgecolors='darkred', linewidth=1.5)
        
        axes[1].set_title(f"{results['symbol']} - Price & Trading Signals", fontsize=14, fontweight='bold')
        axes[1].set_xlabel('Date', fontsize=12)
        axes[1].set_ylabel('Price (VNĐ)', fontsize=12)
        axes[1].legend(loc='best')
        axes[1].grid(True, alpha=0.3)
        axes[1].yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'{x:,.0f}'))
        
        # Add text with key metrics
        fig.text(0.99, 0.01, 
                f"Return: {results['total_return']:.2f}% | Win Rate: {results['win_rate']:.1f}% | Sharpe: {results['sharpe_ratio']:.2f}",
                ha='right', va='bottom', fontsize=10, style='italic', alpha=0.7)
        
        plt.tight_layout()
        
        # Lưu với timestamp
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'backtest_results/{results["symbol"]}_{timestamp}.png'
            
            print(f"💾 Đang lưu biểu đồ vào: {filename}")
            plt.savefig(filename, dpi=300, bbox_inches='tight')
            
            # Kiểm tra file
            if os.path.exists(filename):
                file_size = os.path.getsize(filename)
                print(f"✅ Biểu đồ đã lưu: {filename} ({file_size} bytes)")
            else:
                print(f"❌ Không tìm thấy file sau khi lưu!")
                
            plt.show()
            
        except Exception as e:
            print(f"❌ Lỗi khi lưu biểu đồ: {e}")
            import traceback
            traceback.print_exc()
    
    def run_multiple_backtest(self, symbols, lookback=500, confidence_threshold=50):
        """Chạy backtest cho nhiều cổ phiếu
        Thêm tham số confidence_threshold để đồng bộ với giao diện gọi."""
        all_results = []
        
        for symbol in symbols:
            try:
                result = self.run_backtest(symbol, lookback=lookback, confidence_threshold=confidence_threshold)
                all_results.append(result)
            except Exception as e:
                print(f"❌ Lỗi backtest {symbol}: {e}")
        
        # Summary
        self._print_summary(all_results)
        
        return all_results
    
    def _print_summary(self, all_results):
        """Tổng kết backtest nhiều cổ phiếu"""
        import os
        from datetime import datetime
        
        print(f"\n{'='*60}")
        print(f"📊 TỔNG KẾT BACKTEST")
        print(f"{'='*60}")
        
        summary_df = pd.DataFrame([{
            'Symbol': r['symbol'],
            'Return (%)': f"{r['total_return']:.2f}",
            'Buy&Hold (%)': f"{r['buy_hold_return']:.2f}",
            'Trades': r['total_trades'],
            'Win Rate (%)': f"{r['win_rate']:.2f}",
            'Sharpe': f"{r['sharpe_ratio']:.2f}"
        } for r in all_results])
        
        print(summary_df.to_string(index=False))
        print(f"{'='*60}\n")
        
        # Tự động lưu summary với Excel
        try:
            os.makedirs('backtest_results', exist_ok=True)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            
            # DataFrame chi tiết cho export
            detailed_summary = pd.DataFrame([{
                'Symbol': r['symbol'],
                'Initial Capital': r['initial_capital'],
                'Final Capital': r['final_capital'],
                'Return (%)': r['total_return'],
                'Buy&Hold (%)': r['buy_hold_return'],
                'Outperformance (%)': r['total_return'] - r['buy_hold_return'],
                'Total Trades': r['total_trades'],
                'Winning Trades': r['winning_trades'],
                'Losing Trades': r['losing_trades'],
                'Win Rate (%)': r['win_rate'],
                'Max Drawdown (%)': r['max_drawdown'],
                'Sharpe Ratio': r['sharpe_ratio']
            } for r in all_results])
            
            # Export Excel với format đẹp
            excel_filename = f'backtest_results/summary_{timestamp}.xlsx'
            
            with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
                detailed_summary.to_excel(writer, sheet_name='Summary', index=False)
                
                # Format worksheet
                worksheet = writer.sheets['Summary']
                
                # Auto-adjust column width
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Header style
                from openpyxl.styles import Font, PatternFill, Alignment
                header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                header_font = Font(color='FFFFFF', bold=True)
                
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Number format
                for row in worksheet.iter_rows(min_row=2):
                    # Return, Buy&Hold, Outperformance columns
                    for idx in [3, 4, 5, 9, 10]:
                        row[idx].number_format = '0.00'
                    # Capital columns
                    for idx in [1, 2]:
                        row[idx].number_format = '#,##0'
                
                # Conditional formatting cho Return
                from openpyxl.formatting.rule import CellIsRule
                green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                
                worksheet.conditional_formatting.add(
                    f'D2:D{len(detailed_summary)+1}',
                    CellIsRule(operator='greaterThan', formula=['0'], fill=green_fill)
                )
                worksheet.conditional_formatting.add(
                    f'D2:D{len(detailed_summary)+1}',
                    CellIsRule(operator='lessThan', formula=['0'], fill=red_fill)
                )
            
            if os.path.exists(excel_filename):
                file_size = os.path.getsize(excel_filename)
                print(f"✅ Đã xuất Excel: {excel_filename} ({file_size:,} bytes)")
            
            # Cũng export CSV đơn giản
            csv_filename = f'backtest_results/summary_{timestamp}.csv'
            detailed_summary.to_csv(csv_filename, index=False, encoding='utf-8-sig')
            print(f"✅ Đã xuất CSV: {csv_filename}\n")
                
        except ImportError:
            # Fallback nếu không có openpyxl
            print("⚠️ Cài đặt openpyxl để xuất Excel đẹp: pip install openpyxl")
            csv_filename = f'backtest_results/summary_{timestamp}.csv'
            detailed_summary.to_csv(csv_filename, index=False, encoding='utf-8-sig')
            print(f"✅ Đã xuất CSV: {csv_filename}\n")
            
        except Exception as e:
            print(f"❌ Lỗi khi xuất file: {e}\n")
            import traceback
            traceback.print_exc()